# ====================================================================================================================

import tkinter as tk
from tkinter import filedialog,messagebox,Label
from tkinter import ttk
import os
from openpyxl import load_workbook,Workbook
import time

# ====================================================================================================================

All_values=[]
errors=[]

# ====================================================================================================================

def read_excel_files(folder_path,file_extension,sheet_name,row_start,row_end,column_start,column_end,progress_bar):
    processed_files=0
    is_first_file=True
    for roots, dirs, files in os.walk(folder_path):
        for filename in files:
            if filename.endswith(file_extension):
                file_path = os.path.join(roots, filename)

                try:
                    wb = load_workbook(file_path,data_only=True)
                    ws=wb.active

                    if sheet_name in wb.sheetnames:
                        sheet = wb[sheet_name]
                        
                        for i in range(row_start, row_end+1):
                            row_values=[]
                            row_values.append(file_path)
                            for j in range(column_start, column_end+1):
                                row_values.append(sheet.cell(row=i, column=j).value)
                            All_values.append(row_values)
                        

                        if is_first_file:
                            is_first_file=False
                            row_start=row_start+1
                            
                        
                        All_values[0][0]="File Path"
                        print(All_values)
                        
                    else:
                        errors.append([file_path, "Sheet1 not found in the excel file"])
                        print(f"File: {file_path} does not contain Sheet1.")

                except Exception as e:
                    print(f"Could not read file {file_path}. Error: {e}")

                processed_files +=1
                progress_bar['value']=processed_files
                root.update_idletasks()

                time.sleep(0.2)
# ====================================================================================================================

def save_results_to_excel(save_path, file_name):

    wb = Workbook()

    data_ws = wb.create_sheet('Data')
    for value in All_values:
        data_ws.append(value)


    error_wb = wb.create_sheet('Error')
    for error in errors:
        error_wb.append(error)

    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    full_path = os.path.join(save_path, f"{file_name}.xlsx")
    wb.save(full_path)
    time.sleep(0.5)
# ====================================================================================================================

def input_browse():
    input_folder_path = filedialog.askdirectory()

    if input_folder_path:
        Input_filepath_var.set(input_folder_path)
    else:
        print("No folder selected.")

def output_browse():
    ouput_folder_path = filedialog.askdirectory()

    if ouput_folder_path:
        Output_filepath_var.set(ouput_folder_path)
    else:
        print("No folder selected.")

# ====================================================================================================================

def submit():
    Input_filepath= Input_filepath_var.get()
    file_extension=file_extension_var.get()
    sheet_name=sheet_name_var.get()
    row_start=row_start_var.get()
    row_end=row_end_var.get()
    column_start=column_start_var.get()
    column_end=column_end_var.get()
    Output_filepath= Output_filepath_var.get()
    file_name = file_name_var.get()

    if not Input_filepath:
        messagebox.showwarning("Warning", "Input File Path location should not be Empty.")
        return
    if not file_extension:
        messagebox.showwarning("Warning", "File Extension should not be Empty.")
        return
    if not sheet_name:
        messagebox.showwarning("Warning", "Worksheet name should not be Empty.")
        return
    if not row_start:
        messagebox.showwarning("Warning", "Start Row should not be Empty.")
        return
    if not row_end:
        messagebox.showwarning("Warning", "End Row should not be Empty.")
        return
    if not column_start:
        messagebox.showwarning("Warning", "Start Column should not be Empty.")
        return
    if not column_end:
        messagebox.showwarning("Warning", "End Column should not be Empty.")
        return
    if not Output_filepath:
        messagebox.showwarning("Warning", "Output File Path Location should not be Empty.")
        return
    if not file_name:
        messagebox.showwarning("Warning", "Name of Output Excel File should not be Empty.")
        return

    total_files = 0
    # os.walk generates the file names in a directory tree
    for r, d, files in os.walk(Input_filepath):
        total_files += len(files) # Count the number of files in the current directory

    Answer = messagebox.askyesno("Confirmation", f" Total number of Files: {total_files} !   Do you want to Proceed ?")

    if Answer:
        progress_window = tk.Toplevel(root)
        progress_window.title("Processing Files")
        progress_window.geometry("400x100")
        progress_window.transient(root)
        progress_window.grab_set()
        progress_window.update_idletasks()

        x = root.winfo_x() + (root.winfo_width() // 2) - (progress_window.winfo_reqwidth() // 2)
        y = root.winfo_y() + (root.winfo_height() // 2) - (progress_window.winfo_reqheight() // 2)
        progress_window.geometry(f"+{x}+{y}")

        tk.Label(progress_window, text="Processing... Please wait").pack(pady=10)
        progress_bar = ttk.Progressbar(progress_window, orient="horizontal", length=350, mode="determinate",
                                       maximum=total_files)
        progress_bar.pack(pady=10)

        read_excel_files(Input_filepath,file_extension,sheet_name,row_start,row_end,column_start,column_end,progress_bar)
        save_results_to_excel(Output_filepath, file_name)

        progress_bar.place_forget()
        progress_window.destroy()

        messagebox.showinfo("Success", f"All Excel files are imported successfully at {Output_filepath} and File Name is {file_name} .xlsx")
        All_values.clear()
        errors.clear()

# ====================================================================================================================
def exiit():
    root.destroy()
# ====================================================================================================================

root = tk.Tk()
root.title("Join Excel Sheet Application")
root.geometry("1000x700")


Input_filepath_var= tk.StringVar()
file_extension_var=tk.StringVar(value=".xlsx")
sheet_name_var=tk.StringVar()
row_start_var=tk.IntVar()
row_end_var=tk.IntVar()
column_start_var=tk.IntVar()
column_end_var=tk.IntVar()
Output_filepath_var= tk.StringVar()
file_name_var = tk.StringVar()

# ====================================================================================================================
# ====================================================================================================================
tk.Label(root, text="Join Excel Sheets ", fg="blue",font=("arial", 20)).place(x=400,y=20)
# ====================================================================================================================
tk.Label(root, text="1. Select the Input File Path Location :", font=("arial", 14)).place(x=10,y=70)
Input_filepath_entry = tk.Entry(root, textvariable=Input_filepath_var, font=("arial", 14)).place(x=400,y=78)
Input_path_button = tk.Button(root, text=" Browse ", command=input_browse, font=("arial", 14)).place(x=650,y=70)
# ====================================================================================================================
tk.Label(root, text="2. Input Excel File Extension: ", font=("arial", 14)).place(x=10,y=120)
file_extension_entry = tk.Entry(root, textvariable=file_extension_var, font=("arial", 14)).place(x=400,y=120)
tk.Label(root, text="Example: .xlsx / .xls   / .xlsm ", font=("arial", 10)).place(x=650,y=120)

# ====================================================================================================================
tk.Label(root, text="3. Enter the Name of worksheet to Read:", font=("arial", 14)).place(x=10,y=170)
file_extension_entry = tk.Entry(root, textvariable=sheet_name_var, font=("arial", 14)).place(x=400, y=170)
tk.Label(root, text="Example: Sheet1 / Sheet2 / etc ", font=("arial", 10)).place(x=650,y=170)
# ====================================================================================================================
tk.Label(root, text="4. Enter the Starting Row Number :", font=("arial", 14)).place(x=10,y=220)
row_start_entry  = tk.Entry(root, textvariable=row_start_var, font=("arial", 14)).place(x=400, y=220)
tk.Label(root, text="Example: First Row means , Enter 1  ", font=("arial", 10)).place(x=650,y=220)

# ====================================================================================================================
tk.Label(root, text="5. Enter the Ending Row Number :", font=("arial", 14)).place(x=10,y=270)
row_end_entry   = tk.Entry(root, textvariable=row_end_var, font=("arial", 14)).place(x=400, y=270)
tk.Label(root, text="Example: Sixth Row means , Enter 6  ", font=("arial", 10)).place(x=650,y=270)

# ====================================================================================================================
tk.Label(root, text="6. Enter the Starting Column Number :", font=("arial", 14)).place(x=10,y=320)
column_start_entry  = tk.Entry(root, textvariable=column_start_var, font=("arial", 14)).place(x=400, y=320)
tk.Label(root, text="Example: A Column means , Enter 1  ", font=("arial", 10)).place(x=650,y=320)

# ====================================================================================================================

tk.Label(root, text="7. Enter the Ending Column Number :", font=("arial", 14)).place(x=10,y=370)
column_end_entry  = tk.Entry(root, textvariable=column_end_var, font=("arial", 14)).place(x=400, y=370)
tk.Label(root, text="Example: C Column means , Enter 3  ", font=("arial", 10)).place(x=650,y=370)

# ====================================================================================================================

tk.Label(root, text="8. Select the Output File Path Location :", font=("arial", 14)).place(x=10,y=420)
output_filepath_entry = tk.Entry(root, textvariable=Output_filepath_var, font=("arial", 14)).place(x=400,y=428)
Output_path_button = tk.Button(root, text=" Browse ", command=output_browse, font=("arial", 14)).place(x=650,y=420)
# ====================================================================================================================

tk.Label(root, text="9. Enter the Name of Output Excel File :", font=("arial", 14)).place(x=10,y=470)
file_name_entry = tk.Entry(root, textvariable=file_name_var, font=("arial", 14)).place(x=400, y=470)
tk.Label(root, text="Example: FinalData  ", font=("arial", 10)).place(x=650,y=470)

# ====================================================================================================================

submit_button = tk.Button(root, text=" Process", command=submit, font=("arial", 14, "bold")).place(x=330,y=520)

exit_button = tk.Button(root, text="Exit", command=exiit, font=("arial", 14, "bold")).place(x=530,y=520)

root.mainloop()
# ====================================================================================================================
